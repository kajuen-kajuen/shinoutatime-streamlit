"""
Excel Repositoryのユニットテスト

要件2.1, 2.2, 2.3, 2.4, 8.1, 8.2, 8.3, 8.4, 8.5をテスト
"""

import tempfile
from pathlib import Path
import pytest
from openpyxl import Workbook

from src.repositories.excel_repository import ExcelRepository
from src.exceptions.errors import DataLoadError


class TestExcelRepositorySheetNames:
    """シート名取得のテスト（要件2.1, 2.2）"""
    
    def test_get_sheet_names_success(self):
        """正常なシート名取得"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # テスト用Excelファイルを作成
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            wb.create_sheet("Sheet1")
            wb.create_sheet("Sheet2")
            wb.remove(wb["Sheet"])  # デフォルトシートを削除
            wb.save(excel_path)
            
            # リポジトリを作成してシート名を取得
            repo = ExcelRepository(str(excel_path))
            sheet_names = repo.get_sheet_names()
            
            assert "Sheet1" in sheet_names
            assert "Sheet2" in sheet_names
            assert len(sheet_names) == 2
    
    def test_get_sheet_names_file_not_exists(self):
        """存在しないファイルの処理（要件8.1）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "nonexistent.xlsx"
            
            repo = ExcelRepository(str(excel_path))
            
            with pytest.raises(DataLoadError) as exc_info:
                repo.get_sheet_names()
            
            assert "ファイルが存在しません" in str(exc_info.value)


class TestExcelRepositorySheetExists:
    """シート存在確認のテスト（要件2.3, 2.4）"""
    
    def test_sheet_exists_true(self):
        """シートが存在する場合"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            wb.create_sheet("M_YT_LIVE")
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            
            assert repo.sheet_exists("M_YT_LIVE") is True
    
    def test_sheet_exists_false(self):
        """シートが存在しない場合"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            
            assert repo.sheet_exists("NonExistent") is False
    
    def test_sheet_exists_case_insensitive(self):
        """大文字小文字を区別しない（要件2.4）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            wb.create_sheet("M_YT_LIVE")
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            
            # 小文字でも検索できる
            assert repo.sheet_exists("m_yt_live") is True
            # 大文字でも検索できる
            assert repo.sheet_exists("M_YT_LIVE") is True
            # 混在でも検索できる
            assert repo.sheet_exists("M_yt_LiVe") is True


class TestExcelRepositoryLoadSheet:
    """シートデータ読み込みのテスト（要件2.1, 2.2）"""
    
    def test_load_sheet_success(self):
        """正常なシートデータ読み込み"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            ws = wb.create_sheet("TestSheet")
            
            # テストデータを追加
            ws.append(["ID", "Name", "Value"])
            ws.append([1, "Test1", 100])
            ws.append([2, "Test2", 200])
            
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            rows = repo.load_sheet("TestSheet")
            
            assert len(rows) == 3
            assert rows[0] == ["ID", "Name", "Value"]
            assert rows[1] == [1, "Test1", 100]
            assert rows[2] == [2, "Test2", 200]
    
    def test_load_sheet_case_insensitive(self):
        """大文字小文字を区別しない読み込み（要件2.4）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            ws = wb.create_sheet("M_YT_LIVE")
            ws.append(["ID", "Title"])
            ws.append([1, "Test"])
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            
            # 小文字で読み込み
            rows = repo.load_sheet("m_yt_live")
            assert len(rows) == 2
    
    def test_load_sheet_not_exists(self):
        """存在しないシートの読み込み（要件2.3）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            
            with pytest.raises(DataLoadError) as exc_info:
                repo.load_sheet("NonExistent")
            
            assert "シートが見つかりません" in str(exc_info.value)
    
    def test_load_sheet_empty(self):
        """空のシートの読み込み（要件8.5）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            ws = wb.create_sheet("EmptySheet")
            # 空のシート（データなし）
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            rows = repo.load_sheet("EmptySheet")
            
            # 空行はスキップされるので、空のリストが返される
            assert len(rows) == 0
    
    def test_load_sheet_with_unicode(self):
        """Unicode文字を含むシートの読み込み"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            ws = wb.create_sheet("TestSheet")
            
            # 日本語データを追加
            ws.append(["ID", "タイトル", "アーティスト"])
            ws.append([1, "かすかなおと", "幽音しの"])
            ws.append([2, "テスト曲", "テストアーティスト"])
            
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            rows = repo.load_sheet("TestSheet")
            
            assert len(rows) == 3
            assert rows[0] == ["ID", "タイトル", "アーティスト"]
            assert rows[1] == [1, "かすかなおと", "幽音しの"]
    
    def test_load_sheet_skips_empty_rows(self):
        """空行をスキップする"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            ws = wb.create_sheet("TestSheet")
            
            # データと空行を追加
            ws.append(["ID", "Name"])
            ws.append([1, "Test1"])
            ws.append([None, None])  # 空行
            ws.append([2, "Test2"])
            ws.append(["", ""])  # 空文字列の行
            ws.append([3, "Test3"])
            
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            rows = repo.load_sheet("TestSheet")
            
            # 空行はスキップされる
            assert len(rows) == 4
            assert rows[0] == ["ID", "Name"]
            assert rows[1] == [1, "Test1"]
            assert rows[2] == [2, "Test2"]
            assert rows[3] == [3, "Test3"]


class TestExcelRepositoryErrors:
    """エラーハンドリングのテスト（要件8.1, 8.2, 8.3, 8.4）"""
    
    def test_file_not_exists(self):
        """ファイルが存在しない（要件8.1）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "nonexistent.xlsx"
            
            repo = ExcelRepository(str(excel_path))
            
            with pytest.raises(DataLoadError) as exc_info:
                repo.load_sheet("AnySheet")
            
            assert "ファイルが存在しません" in str(exc_info.value)
    
    def test_invalid_file_format(self):
        """不正なファイル形式（要件8.4）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "invalid.xlsx"
            
            # テキストファイルを.xlsx拡張子で保存
            excel_path.write_text("This is not an Excel file", encoding='utf-8')
            
            repo = ExcelRepository(str(excel_path))
            
            with pytest.raises(DataLoadError) as exc_info:
                repo.load_sheet("AnySheet")
            
            assert "読み込みに失敗しました" in str(exc_info.value)
            assert "破損しているか、形式が不正です" in str(exc_info.value)


class TestExcelRepositoryClose:
    """リソース管理のテスト"""
    
    def test_close_workbook(self):
        """ワークブックを閉じる"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            wb.create_sheet("TestSheet")
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            
            # ワークブックを読み込む
            repo.get_sheet_names()
            
            # ワークブックを閉じる
            repo.close()
            
            # 閉じた後は再度読み込みが必要
            assert repo._workbook is None
    
    def test_close_without_loading(self):
        """読み込み前にcloseを呼んでもエラーにならない"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            wb.save(excel_path)
            
            repo = ExcelRepository(str(excel_path))
            
            # 読み込み前にcloseを呼ぶ
            repo.close()  # エラーにならないことを確認
            
            assert repo._workbook is None
