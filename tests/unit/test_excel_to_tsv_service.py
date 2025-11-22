"""
Excel to TSV Serviceのユニットテスト

要件1.1, 1.2, 2.1, 2.2, 2.3, 2.4, 2.5をテスト
"""

import tempfile
from pathlib import Path
import pytest
from openpyxl import Workbook

from src.services.excel_to_tsv_service import ExcelToTsvService
from src.repositories.excel_repository import ExcelRepository
from src.repositories.tsv_repository import TsvRepository
from src.repositories.backup_repository import BackupRepository


class TestExcelToTsvServiceConversion:
    """変換処理のテスト"""
    
    def test_convert_excel_to_tsv_success(self):
        """正常な変換処理（要件1.1, 1.2）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # テスト用Excelファイルを作成
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            
            # M_YT_LIVEシートを作成
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            ws_live.append([1, "2024-01-01", "テスト配信", "https://example.com/1"])
            ws_live.append([2, "2024-01-02", "テスト配信2", "https://example.com/2"])
            
            # M_YT_LIVE_TIMESTAMPシートを作成
            ws_timestamp = wb.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp.append([1, 1, "00:00:00", "曲1", "アーティスト1"])
            ws_timestamp.append([2, 1, "00:05:00", "曲2", "アーティスト2"])
            
            wb.remove(wb["Sheet"])  # デフォルトシートを削除
            wb.save(excel_path)
            
            # サービスを作成
            output_dir = Path(tmpdir) / "output"
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # 変換を実行
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            # 結果を検証
            assert result.success is True
            assert len(result.files_created) == 2
            assert len(result.errors) == 0
            
            # ファイルが作成されたことを確認
            live_file = output_dir / "M_YT_LIVE.TSV"
            timestamp_file = output_dir / "M_YT_LIVE_TIMESTAMP.TSV"
            
            assert live_file.exists()
            assert timestamp_file.exists()
            
            # ファイルの内容を確認
            live_content = live_file.read_text(encoding='utf-8')
            assert "ID\t配信日\tタイトル\tURL" in live_content
            assert "1\t2024-01-01\tテスト配信\thttps://example.com/1" in live_content
    
    def test_convert_excel_to_tsv_file_not_exists(self):
        """存在しないファイルの処理（要件1.3）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "nonexistent.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            assert result.success is False
            assert len(result.errors) > 0
            assert "ファイルが存在しません" in result.errors[0]
    
    def test_convert_excel_to_tsv_missing_sheets(self):
        """必要なシートが存在しない場合（要件2.5）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # M_YT_LIVEシートのみ作成（M_YT_LIVE_TIMESTAMPがない）
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            output_dir = Path(tmpdir) / "output"
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            assert result.success is False
            assert len(result.errors) > 0
            assert "必要なシートが存在しません" in result.errors[0]
    
    def test_convert_excel_to_tsv_dry_run(self):
        """ドライランモード（要件6.5）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # テスト用Excelファイルを作成
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            ws_live.append([1, "2024-01-01", "テスト", "https://example.com"])
            
            ws_timestamp = wb.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp.append([1, 1, "00:00:00", "曲", "アーティスト"])
            
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            output_dir = Path(tmpdir) / "output"
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # ドライランで実行
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=True
            )
            
            # 結果を検証
            assert result.success is True
            assert len(result.files_created) == 0  # ドライランではファイルを作成しない
            assert len(result.errors) == 0
            
            # ファイルが作成されていないことを確認
            live_file = output_dir / "M_YT_LIVE.TSV"
            timestamp_file = output_dir / "M_YT_LIVE_TIMESTAMP.TSV"
            
            assert not live_file.exists()
            assert not timestamp_file.exists()
    
    def test_convert_excel_to_tsv_with_backup(self):
        """既存ファイルのバックアップ（要件1.5, 7.1）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # テスト用Excelファイルを作成
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            ws_live.append([1, "2024-01-01", "新しいデータ", "https://example.com"])
            
            ws_timestamp = wb.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp.append([1, 1, "00:00:00", "曲", "アーティスト"])
            
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            output_dir = Path(tmpdir) / "output"
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # 既存のTSVファイルを作成
            existing_file = output_dir / "M_YT_LIVE.TSV"
            existing_file.write_text("古いデータ", encoding='utf-8')
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # 変換を実行
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            # 結果を検証
            assert result.success is True
            assert len(result.backup_files) > 0  # バックアップが作成された
            
            # バックアップファイルが存在することを確認
            backup_dir = output_dir / "backups"
            assert backup_dir.exists()
            backup_files = list(backup_dir.glob("M_YT_LIVE_*.TSV"))
            assert len(backup_files) > 0
    
    def test_convert_excel_to_tsv_empty_sheet(self):
        """空のシートの処理（要件8.5）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            
            # 空のシートを作成（ヘッダーのみ）
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            # データ行なし
            
            ws_timestamp = wb.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp.append([1, 1, "00:00:00", "曲", "アーティスト"])
            
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            output_dir = Path(tmpdir) / "output"
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            # 空のシートは警告が出るがスキップされる
            assert len(result.warnings) > 0
            # M_YT_LIVE_TIMESTAMPのみ作成される
            assert len(result.files_created) == 1


class TestExcelToTsvServiceValidation:
    """データ検証のテスト"""
    
    def test_validate_sheet_data_field_count(self):
        """フィールド数の検証（要件4.1, 4.2）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # 正しいフィールド数
            rows = [
                [1, "2024-01-01", "タイトル", "https://example.com"],
                [2, "2024-01-02", "タイトル2", "https://example.com/2"]
            ]
            warnings = service.validate_sheet_data("M_YT_LIVE", rows, 4)
            assert len(warnings) == 0
            
            # 不正なフィールド数
            rows_invalid = [
                [1, "2024-01-01", "タイトル"],  # 3フィールド（4が期待される）
            ]
            warnings = service.validate_sheet_data("M_YT_LIVE", rows_invalid, 4)
            assert len(warnings) > 0
            assert "フィールド数が不正です" in warnings[0].message
    
    def test_validate_sheet_data_empty_fields(self):
        """空フィールドの検証（要件4.3）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # 空フィールドを含む行
            rows = [
                [1, "", "タイトル", "https://example.com"],  # 配信日が空
                [2, "2024-01-02", None, "https://example.com/2"]  # タイトルがNone
            ]
            warnings = service.validate_sheet_data("M_YT_LIVE", rows, 4)
            
            # 空フィールドの警告が出る
            empty_warnings = [w for w in warnings if "必須フィールドが空です" in w.message]
            assert len(empty_warnings) > 0
    
    def test_validate_sheet_data_invalid_id(self):
        """IDフィールドの検証（要件4.5）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # 不正なID（数値でない）
            rows = [
                ["abc", "2024-01-01", "タイトル", "https://example.com"],
            ]
            warnings = service.validate_sheet_data("M_YT_LIVE", rows, 4)
            
            # IDが数値でない警告が出る
            id_warnings = [w for w in warnings if "IDフィールドが数値ではありません" in w.message]
            assert len(id_warnings) > 0
    
    def test_validate_sheet_data_invalid_url(self):
        """URL形式の検証（要件4.4）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # 不正なURL
            rows = [
                [1, "2024-01-01", "タイトル", "not-a-url"],
            ]
            warnings = service.validate_sheet_data("M_YT_LIVE", rows, 4)
            
            # URL形式が不正の警告が出る
            url_warnings = [w for w in warnings if "URL形式が不正です" in w.message]
            assert len(url_warnings) > 0


class TestExcelToTsvServiceURLValidation:
    """URL検証のテスト"""
    
    def test_is_valid_url_valid_urls(self):
        """有効なURL"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # 有効なURL
            assert service._is_valid_url("https://example.com") is True
            assert service._is_valid_url("http://example.com") is True
            assert service._is_valid_url("https://example.com/path") is True
            assert service._is_valid_url("https://example.com:8080") is True
    
    def test_is_valid_url_invalid_urls(self):
        """無効なURL"""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # 無効なURL
            assert service._is_valid_url("not-a-url") is False
            assert service._is_valid_url("ftp://example.com") is False
            assert service._is_valid_url("example.com") is False
            assert service._is_valid_url("") is False


class TestExcelToTsvServiceSongListGenerator:
    """song_list_generator実行のテスト"""
    
    @pytest.mark.skip(reason="song_list_generatorモジュールが環境にインストールされていない場合はスキップ")
    def test_run_song_list_generator_success(self):
        """song_list_generatorの正常実行（要件9.1, 9.2, 9.3）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            
            # テスト用のTSVファイルを作成
            live_file = tmpdir_path / "M_YT_LIVE.TSV"
            timestamp_file = tmpdir_path / "M_YT_LIVE_TIMESTAMP.TSV"
            output_file = tmpdir_path / "V_SONG_LIST.TSV"
            
            # M_YT_LIVE.TSVを作成（YYYY/M/D形式の日付）
            live_file.write_text(
                "ID\t配信日\tタイトル\tURL\n"
                "1\t2024/1/1\tテスト配信\thttps://example.com/1\n",
                encoding='utf-8'
            )
            
            # M_YT_LIVE_TIMESTAMP.TSVを作成
            timestamp_file.write_text(
                "ID\tLIVE_ID\tタイムスタンプ\t曲名\tアーティスト\n"
                "1\t1\t00:00:00\tテスト曲\tテストアーティスト\n",
                encoding='utf-8'
            )
            
            # サービスを作成
            excel_path = tmpdir_path / "test.xlsx"
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(tmpdir_path))
            backup_repo = BackupRepository(str(tmpdir_path / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # song_list_generatorを実行
            result = service.run_song_list_generator(
                str(live_file),
                str(timestamp_file),
                str(output_file)
            )
            
            # 結果を検証
            assert result is True
            # V_SONG_LIST.TSVが生成されたことを確認
            assert output_file.exists()
    
    def test_run_song_list_generator_missing_files(self):
        """存在しないファイルでの実行（要件9.4）"""
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            
            # 存在しないファイルパス
            live_file = tmpdir_path / "nonexistent_live.tsv"
            timestamp_file = tmpdir_path / "nonexistent_timestamp.tsv"
            output_file = tmpdir_path / "output.tsv"
            
            # サービスを作成
            excel_path = tmpdir_path / "test.xlsx"
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(tmpdir_path))
            backup_repo = BackupRepository(str(tmpdir_path / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # song_list_generatorを実行（失敗するはず）
            result = service.run_song_list_generator(
                str(live_file),
                str(timestamp_file),
                str(output_file)
            )
            
            # 結果を検証（失敗するがエラーは発生しない）
            assert result is False
