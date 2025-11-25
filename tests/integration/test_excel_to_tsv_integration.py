"""
Excel to TSV変換の統合テスト

エンドツーエンドの変換処理をテストします。
全要件を統合的に検証します。
"""

import tempfile
from pathlib import Path
import pytest
from openpyxl import Workbook
import subprocess
import sys

from src.services.excel_to_tsv_service import ExcelToTsvService
from src.repositories.excel_repository import ExcelRepository
from src.repositories.tsv_repository import TsvRepository
from src.repositories.backup_repository import BackupRepository


class TestExcelToTsvIntegrationEndToEnd:
    """正常系のエンドツーエンドテスト"""
    
    def test_full_conversion_workflow(self):
        """
        完全な変換ワークフローのテスト
        
        要件: 1.1, 1.2, 2.1, 2.2, 3.1, 3.2, 3.5, 4.1, 4.2
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            # テスト用Excelファイルを作成
            excel_path = Path(tmpdir) / "test_data.xlsx"
            wb = Workbook()
            
            # M_YT_LIVEシートを作成
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            ws_live.append([1, "2024-01-01", "新年配信", "https://youtube.com/watch?v=1"])
            ws_live.append([2, "2024-01-15", "歌枠配信", "https://youtube.com/watch?v=2"])
            ws_live.append([3, "2024-02-01", "雑談配信", "https://youtube.com/watch?v=3"])
            
            # M_YT_LIVE_TIMESTAMPシートを作成
            ws_timestamp = wb.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp.append([1, 1, "00:05:30", "曲A", "アーティストX"])
            ws_timestamp.append([2, 1, "00:15:20", "曲B", "アーティストY"])
            ws_timestamp.append([3, 2, "00:10:00", "曲C", "アーティストZ"])
            ws_timestamp.append([4, 2, "00:25:45", "曲D", "アーティストW"])
            
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            # 出力ディレクトリを設定
            output_dir = Path(tmpdir) / "output"
            
            # サービスを作成して変換を実行
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            # 結果を検証
            assert result.success is True
            assert len(result.files_created) == 2
            assert len(result.errors) == 0
            
            # 生成されたファイルを検証
            live_file = output_dir / "M_YT_LIVE.TSV"
            timestamp_file = output_dir / "M_YT_LIVE_TIMESTAMP.TSV"
            
            assert live_file.exists()
            assert timestamp_file.exists()
            
            # M_YT_LIVE.TSVの内容を検証
            live_content = live_file.read_text(encoding='utf-8')
            lines = live_content.strip().split('\n')
            
            # ヘッダー行を検証（要件3.5）
            assert lines[0] == "ID\t配信日\tタイトル\tURL"
            
            # データ行を検証（要件3.1, 4.1）
            assert len(lines) == 4  # ヘッダー + 3データ行
            # Excelから読み込んだ日付は YYYY/M/D 形式で出力される
            assert "1\t2024/1/1\t新年配信\thttps://youtube.com/watch?v=1" in lines[1]
            assert "2\t2024/1/15\t歌枠配信\thttps://youtube.com/watch?v=2" in lines[2]
            assert "3\t2024/2/1\t雑談配信\thttps://youtube.com/watch?v=3" in lines[3]
            
            # M_YT_LIVE_TIMESTAMP.TSVの内容を検証
            timestamp_content = timestamp_file.read_text(encoding='utf-8')
            timestamp_lines = timestamp_content.strip().split('\n')
            
            # ヘッダー行を検証
            assert timestamp_lines[0] == "ID\tLIVE_ID\tタイムスタンプ\t曲名\tアーティスト"
            
            # データ行を検証（要件4.2）
            assert len(timestamp_lines) == 5  # ヘッダー + 4データ行
            assert "1\t1\t00:05:30\t曲A\tアーティストX" in timestamp_lines[1]
            assert "4\t2\t00:25:45\t曲D\tアーティストW" in timestamp_lines[4]
    
    def test_conversion_with_special_characters(self):
        """
        特殊文字を含むデータの変換テスト
        
        要件: 3.3, 3.4
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test_special.xlsx"
            wb = Workbook()
            
            # 特殊文字を含むデータ
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            ws_live.append([1, "2024-01-01", "タイトル\nに改行", "https://example.com/1"])
            ws_live.append([2, "2024-01-02", "タイトル\tにタブ", "https://example.com/2"])
            
            ws_timestamp = wb.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp.append([1, 1, "00:00:00", "曲名", "アーティスト"])
            
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            assert result.success is True
            
            # 生成されたファイルの内容を検証
            live_file = output_dir / "M_YT_LIVE.TSV"
            live_content = live_file.read_text(encoding='utf-8')
            
            # 改行文字がスペースに置換されていることを確認（要件3.3）
            assert "タイトル に改行" in live_content
            assert "\n" not in live_content.split('\n')[1]  # データ行内に改行がない
            
            # タブ文字がスペースに置換されていることを確認（要件3.4）
            assert "タイトル にタブ" in live_content


class TestExcelToTsvIntegrationBackup:
    """バックアップ機能の統合テスト"""
    
    def test_backup_creation_on_overwrite(self):
        """
        既存ファイル上書き時のバックアップ作成テスト
        
        要件: 1.5, 7.1, 7.2, 7.3, 7.5
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            # 最初の変換
            excel_path = Path(tmpdir) / "test.xlsx"
            wb = Workbook()
            
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            ws_live.append([1, "2024-01-01", "古いデータ", "https://example.com/1"])
            
            ws_timestamp = wb.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp.append([1, 1, "00:00:00", "曲", "アーティスト"])
            
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            output_dir = Path(tmpdir) / "output"
            
            # 最初の変換を実行
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            result1 = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            assert result1.success is True
            assert len(result1.backup_files) == 0  # 最初は既存ファイルがないのでバックアップなし
            
            # 既存ファイルの内容を確認
            live_file = output_dir / "M_YT_LIVE.TSV"
            old_content = live_file.read_text(encoding='utf-8')
            assert "古いデータ" in old_content
            
            # 新しいデータで2回目の変換
            wb2 = Workbook()
            ws_live2 = wb2.create_sheet("M_YT_LIVE")
            ws_live2.append(["ID", "配信日", "タイトル", "URL"])
            ws_live2.append([1, "2024-01-02", "新しいデータ", "https://example.com/2"])
            
            ws_timestamp2 = wb2.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp2.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp2.append([1, 1, "00:00:00", "曲", "アーティスト"])
            
            wb2.remove(wb2["Sheet"])
            wb2.save(excel_path)
            
            # 2回目の変換を実行
            excel_repo2 = ExcelRepository(str(excel_path))
            service2 = ExcelToTsvService(excel_repo2, tsv_repo, backup_repo)
            result2 = service2.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            assert result2.success is True
            assert len(result2.backup_files) > 0  # バックアップが作成された（要件7.1）
            
            # バックアップファイルが存在することを確認
            backup_dir = output_dir / "backups"
            assert backup_dir.exists()  # バックアップディレクトリが作成された（要件7.3）
            
            # M_YT_LIVEのバックアップファイルを検索（タイムスタンプを含むファイル名）
            all_backup_files = list(backup_dir.glob("*.TSV"))
            assert len(all_backup_files) >= 2  # 2つのTSVファイルがバックアップされる
            
            # M_YT_LIVEのバックアップファイルを特定
            live_backup_files = [f for f in all_backup_files if f.name.startswith("M_YT_LIVE_") and "TIMESTAMP" not in f.name]
            assert len(live_backup_files) > 0
            
            # バックアップファイル名にタイムスタンプが含まれることを確認（要件7.2）
            backup_file = live_backup_files[0]
            assert "_" in backup_file.stem  # タイムスタンプ形式を含む
            
            # バックアップファイルの内容が古いデータであることを確認（要件7.5）
            backup_content = backup_file.read_text(encoding='utf-8')
            assert "古いデータ" in backup_content
            
            # 新しいファイルの内容が更新されていることを確認
            new_content = live_file.read_text(encoding='utf-8')
            assert "新しいデータ" in new_content
            assert "古いデータ" not in new_content


class TestExcelToTsvIntegrationErrorRecovery:
    """エラーリカバリの統合テスト"""
    
    def test_missing_required_sheets(self):
        """
        必要なシートが存在しない場合のエラー処理テスト
        
        要件: 2.3, 2.5, 8.4
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            # M_YT_LIVEシートのみ作成（M_YT_LIVE_TIMESTAMPがない）
            excel_path = Path(tmpdir) / "test_incomplete.xlsx"
            wb = Workbook()
            
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            ws_live.append([1, "2024-01-01", "テスト", "https://example.com/1"])
            
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            # エラーが発生することを確認（要件2.5）
            assert result.success is False
            assert len(result.errors) > 0
            assert "必要なシートが存在しません" in result.errors[0]
            
            # ファイルが作成されていないことを確認（要件8.4）
            live_file = output_dir / "M_YT_LIVE.TSV"
            assert not live_file.exists()
    
    def test_file_not_found_error(self):
        """
        ファイルが存在しない場合のエラー処理テスト
        
        要件: 1.3, 8.1
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            # 存在しないファイルパス
            excel_path = Path(tmpdir) / "nonexistent.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            # エラーが発生することを確認（要件1.3）
            assert result.success is False
            assert len(result.errors) > 0
            assert "ファイルが存在しません" in result.errors[0]
    
    def test_empty_sheet_handling(self):
        """
        空のシートの処理テスト
        
        要件: 8.5
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "test_empty.xlsx"
            wb = Workbook()
            
            # 空のシート（ヘッダーのみ）
            ws_live = wb.create_sheet("M_YT_LIVE")
            ws_live.append(["ID", "配信日", "タイトル", "URL"])
            # データ行なし
            
            # 正常なシート
            ws_timestamp = wb.create_sheet("M_YT_LIVE_TIMESTAMP")
            ws_timestamp.append(["ID", "LIVE_ID", "タイムスタンプ", "曲名", "アーティスト"])
            ws_timestamp.append([1, 1, "00:00:00", "曲", "アーティスト"])
            
            wb.remove(wb["Sheet"])
            wb.save(excel_path)
            
            output_dir = Path(tmpdir) / "output"
            
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(output_dir))
            backup_repo = BackupRepository(str(output_dir / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            result = service.convert_excel_to_tsv(
                str(excel_path),
                str(output_dir),
                dry_run=False
            )
            
            # 警告が出るがスキップされる（要件8.5）
            assert len(result.warnings) > 0
            warning_messages = [w.message for w in result.warnings]
            assert any("空です" in msg for msg in warning_messages)
            
            # M_YT_LIVE_TIMESTAMPのみ作成される
            assert len(result.files_created) == 1
            timestamp_file = output_dir / "M_YT_LIVE_TIMESTAMP.TSV"
            assert timestamp_file.exists()


class TestExcelToTsvIntegrationSongListGenerator:
    """後続処理（song_list_generator）の統合テスト"""
    
    def test_song_list_generator_execution(self):
        """
        song_list_generatorの実行テスト
        
        要件: 9.1, 9.2, 9.3
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            
            # テスト用のTSVファイルを作成
            live_file = tmpdir_path / "M_YT_LIVE.TSV"
            timestamp_file = tmpdir_path / "M_YT_LIVE_TIMESTAMP.TSV"
            output_file = tmpdir_path / "V_SONG_LIST.TSV"
            
            # M_YT_LIVE.TSVを作成（YYYY/M/D形式の日付）
            live_file.write_text(
                "ID\t配信日\tタイトル\tURL\n"
                "1\t2024/1/1\t新年配信\thttps://youtube.com/watch?v=1\n"
                "2\t2024/1/15\t歌枠配信\thttps://youtube.com/watch?v=2\n",
                encoding='utf-8'
            )
            
            # M_YT_LIVE_TIMESTAMP.TSVを作成
            timestamp_file.write_text(
                "ID\tLIVE_ID\tタイムスタンプ\t曲名\tアーティスト\n"
                "1\t1\t00:05:30\t曲A\tアーティストX\n"
                "2\t1\t00:15:20\t曲B\tアーティストY\n"
                "3\t2\t00:10:00\t曲C\tアーティストZ\n",
                encoding='utf-8'
            )
            
            # サービスを作成
            excel_path = tmpdir_path / "test.xlsx"
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(tmpdir_path))
            backup_repo = BackupRepository(str(tmpdir_path / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # song_list_generatorを実行（要件9.1, 9.2）
            result = service.run_song_list_generator(
                str(live_file),
                str(timestamp_file),
                str(output_file)
            )
            
            # 結果を検証
            assert result is True
            
            # V_SONG_LIST.TSVが生成されたことを確認（要件9.3）
            assert output_file.exists()
            
            # 生成されたファイルの内容を検証
            output_content = output_file.read_text(encoding='utf-8')
            assert len(output_content) > 0
            
            # ヘッダーが含まれることを確認
            lines = output_content.strip().split('\n')
            assert len(lines) > 0
    
    def test_song_list_generator_error_isolation(self):
        """
        song_list_generatorのエラー分離テスト
        
        要件: 9.4
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            
            # 存在しないファイルでsong_list_generatorを実行
            live_file = tmpdir_path / "nonexistent_live.tsv"
            timestamp_file = tmpdir_path / "nonexistent_timestamp.tsv"
            output_file = tmpdir_path / "output.tsv"
            
            excel_path = tmpdir_path / "test.xlsx"
            excel_repo = ExcelRepository(str(excel_path))
            tsv_repo = TsvRepository(str(tmpdir_path))
            backup_repo = BackupRepository(str(tmpdir_path / "backups"))
            
            service = ExcelToTsvService(excel_repo, tsv_repo, backup_repo)
            
            # song_list_generatorを実行（失敗するはず）
            result = service.run_song_list_generator(
                str(live_file),
                str(timestamp_file),
                str(output_file)
            )
            
            # エラーが発生するが例外は発生しない（要件9.4）
            assert result is False
