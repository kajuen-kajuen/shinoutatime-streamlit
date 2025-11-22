"""
Excelリポジトリモジュール

Excelファイルの読み込みを管理します。
"""

import logging
from pathlib import Path
from typing import Any, List, Optional
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.exceptions.errors import DataLoadError


class ExcelRepository:
    """
    Excelリポジトリ
    
    Excelファイルからシートデータを読み込みます。
    """
    
    def __init__(self, file_path: str):
        """
        リポジトリを初期化
        
        Args:
            file_path: Excelファイルのパス
        """
        self.file_path = Path(file_path)
        self.logger = logging.getLogger(__name__)
        self._workbook: Optional[Workbook] = None
    
    def load_sheet(self, sheet_name: str) -> List[List[Any]]:
        """
        指定されたシートを読み込む
        
        Args:
            sheet_name: シート名（大文字小文字を区別しない）
            
        Returns:
            シートデータ（行のリスト、各行はセル値のリスト）
            
        Raises:
            DataLoadError: シートの読み込みに失敗した場合
        """
        try:
            # ワークブックを読み込む（まだ読み込まれていない場合）
            if self._workbook is None:
                self._load_workbook()
            
            # シートを取得（大文字小文字を区別しない）
            worksheet = self._get_worksheet(sheet_name)
            
            if worksheet is None:
                error_msg = f"シートが見つかりません: {sheet_name}"
                self.logger.error(error_msg)
                raise DataLoadError(
                    file_path=str(self.file_path),
                    message=error_msg
                )
            
            # シートデータを読み込む
            rows = []
            for row in worksheet.iter_rows(values_only=True):
                # 空行をスキップ
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                rows.append(list(row))
            
            self.logger.info(
                f"シートを読み込みました: {sheet_name} ({len(rows)}行)"
            )
            return rows
            
        except DataLoadError:
            raise
        except Exception as e:
            error_msg = f"シートの読み込みに失敗しました: {sheet_name} ({e})"
            self.logger.error(error_msg, exc_info=True)
            raise DataLoadError(
                file_path=str(self.file_path),
                message=error_msg
            )
    
    def get_sheet_names(self) -> List[str]:
        """
        すべてのシート名を取得
        
        Returns:
            シート名のリスト
            
        Raises:
            DataLoadError: ワークブックの読み込みに失敗した場合
        """
        try:
            # ワークブックを読み込む（まだ読み込まれていない場合）
            if self._workbook is None:
                self._load_workbook()
            
            sheet_names = self._workbook.sheetnames
            self.logger.debug(f"シート名を取得しました: {sheet_names}")
            return sheet_names
            
        except Exception as e:
            error_msg = f"シート名の取得に失敗しました: {e}"
            self.logger.error(error_msg, exc_info=True)
            raise DataLoadError(
                file_path=str(self.file_path),
                message=error_msg
            )
    
    def sheet_exists(self, sheet_name: str) -> bool:
        """
        シートが存在するか確認（大文字小文字を区別しない）
        
        Args:
            sheet_name: シート名
            
        Returns:
            シートが存在する場合True
        """
        try:
            # ワークブックを読み込む（まだ読み込まれていない場合）
            if self._workbook is None:
                self._load_workbook()
            
            # 大文字小文字を区別せずに検索
            sheet_name_lower = sheet_name.lower()
            for name in self._workbook.sheetnames:
                if name.lower() == sheet_name_lower:
                    return True
            
            return False
            
        except Exception as e:
            self.logger.warning(
                f"シートの存在確認に失敗しました: {sheet_name} ({e})"
            )
            return False
    
    def _load_workbook(self) -> None:
        """
        Excelワークブックを読み込む
        
        Raises:
            DataLoadError: ワークブックの読み込みに失敗した場合
        """
        # ファイルの存在確認
        if not self.file_path.exists():
            error_msg = f"ファイルが存在しません: {self.file_path}"
            self.logger.error(error_msg)
            raise DataLoadError(
                file_path=str(self.file_path),
                message=error_msg
            )
        
        try:
            # Excelファイルを読み込む
            self._workbook = load_workbook(
                filename=self.file_path,
                read_only=True,  # 読み取り専用モードで高速化
                data_only=True   # 数式ではなく値を取得
            )
            
            self.logger.info(
                f"Excelファイルを読み込みました: {self.file_path}"
            )
            
        except PermissionError as e:
            error_msg = "ファイルへのアクセス権限がありません"
            self.logger.error(f"{error_msg}: {self.file_path}", exc_info=True)
            raise DataLoadError(
                file_path=str(self.file_path),
                message=f"{error_msg}。ファイルが他のプログラムで開かれている可能性があります。"
            ) from e
            
        except Exception as e:
            error_msg = f"Excelファイルの読み込みに失敗しました: {e}"
            self.logger.error(error_msg, exc_info=True)
            raise DataLoadError(
                file_path=str(self.file_path),
                message=f"{error_msg}。ファイルが破損しているか、形式が不正です。"
            )
    
    def _get_worksheet(self, sheet_name: str) -> Optional[Worksheet]:
        """
        シート名からワークシートを取得（大文字小文字を区別しない）
        
        Args:
            sheet_name: シート名
            
        Returns:
            ワークシートオブジェクト（見つからない場合はNone）
        """
        if self._workbook is None:
            return None
        
        # 大文字小文字を区別せずに検索
        sheet_name_lower = sheet_name.lower()
        for name in self._workbook.sheetnames:
            if name.lower() == sheet_name_lower:
                return self._workbook[name]
        
        return None
    
    def close(self) -> None:
        """
        ワークブックを閉じる
        
        リソースを解放します。
        """
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None
            self.logger.debug("ワークブックを閉じました")
